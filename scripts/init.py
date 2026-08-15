#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
交易记录登记助手（simple）初始化脚本。

功能：
1. 在当前项目目录生成「交易记录.xlsx」空模板（含「交易记录」「填写说明」两个 sheet）。
2. 把 skill 内的「使用说明.md」模板复制到当前项目目录。

幂等：若 xlsx 已存在则跳过生成，避免覆盖历史记录。
"""

import os
import shutil
import sys

# 品牌署名，写进 xlsx 填写说明 sheet 底部
BRAND = "公众号：金方土2025"

# 字段定义（照搬《手把手教你用AI做投资研究2》字段说明表）
HEADERS = [
    "交易编号",
    "日期",
    "标的",
    "操作",
    "成交价",
    "数量/金额",
    "仓位比例",
    "交易理由",
    "止损价/风险控制",
    "清仓复盘",
]

FIELD_DESC = [
    ("字段", "怎么填"),
    ("交易编号", "同一轮建仓、加仓、减仓、清仓用同一个编号"),
    ("日期", "实际成交日"),
    ("标的", "股票、ETF、基金等"),
    ("操作", "建仓、加仓、减仓、清仓"),
    ("成交价", "实际价格"),
    ("数量/金额", "二选一即可"),
    ("仓位比例", "占账户资金多少"),
    ("交易理由", "一句话写清楚为什么做"),
    ("止损价/风险控制", "怎样说明原判断错了（价格、基本面）"),
    ("清仓复盘", "卖出原因、是否按计划、下一次注意什么"),
]

XLSX_NAME = "交易记录.xlsx"
MD_NAME = "使用说明.md"


def ensure_openpyxl():
    """确保 openpyxl 可用，缺失时给出明确提示。"""
    try:
        import openpyxl  # noqa: F401
        return openpyxl
    except ImportError:
        print("[错误] 缺少 openpyxl 依赖。请先安装：")
        print("  pip install openpyxl")
        print("  （或让 agent 安装到隔离环境后再运行本脚本）")
        sys.exit(1)


def build_xlsx(target_dir):
    """生成交易记录.xlsx 空模板。"""
    openpyxl = ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Sheet1: 交易记录 ----
    ws = wb.active
    ws.title = "交易记录"

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 列宽（按字段内容合理设置）
    col_widths = [12, 12, 12, 8, 10, 18, 10, 18, 24, 30]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ---- Sheet2: 填写说明 ----
    ws2 = wb.create_sheet("填写说明")
    for r_idx, (field, desc) in enumerate(FIELD_DESC, start=1):
        c1 = ws2.cell(row=r_idx, column=1, value=field)
        c2 = ws2.cell(row=r_idx, column=2, value=desc)
        if r_idx == 1:
            c1.font = header_font
            c1.fill = header_fill
            c2.font = header_font
            c2.fill = header_fill
            c1.alignment = center
            c2.alignment = center

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 42

    # 底部品牌署名
    brand_row = len(FIELD_DESC) + 2
    brand_cell = ws2.cell(row=brand_row, column=1, value=BRAND)
    brand_cell.font = Font(italic=True, color="808080")

    xlsx_path = os.path.join(target_dir, XLSX_NAME)
    wb.save(xlsx_path)
    print(f"[OK] 已生成 {xlsx_path}")
    return xlsx_path


def copy_md(target_dir):
    """把 assets/使用说明.md 复制到目标目录。"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    src = os.path.join(script_dir, "..", "assets", "使用说明.md")
    src = os.path.abspath(src)

    if not os.path.exists(src):
        print(f"[警告] 未找到说明文档模板：{src}，跳过复制。")
        return None

    dst = os.path.join(target_dir, MD_NAME)
    shutil.copyfile(src, dst)
    print(f"[OK] 已生成 {dst}")
    return dst


def main():
    # 目标目录 = 当前工作目录
    target_dir = os.getcwd()

    xlsx_path = os.path.join(target_dir, XLSX_NAME)
    if os.path.exists(xlsx_path):
        print(f"[跳过] {xlsx_path} 已存在，不覆盖。如需重建请先删除或改名。")
    else:
        build_xlsx(target_dir)

    copy_md(target_dir)

    print("\n初始化完成。")
    print(f"请阅读 {os.path.join(target_dir, MD_NAME)} 了解字段填写规范。")


if __name__ == "__main__":
    main()
